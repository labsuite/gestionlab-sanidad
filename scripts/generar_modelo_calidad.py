#!/usr/bin/env python3
"""
generar_modelo_calidad.py
Rellena los dos documentos del modelo de calidad con datos en tiempo real de Supabase:
  1. Inventarios_Sanidade_XXXX_XX.xlsx  — hoja "Sanidade" con todos los equipos
  2. Plan_Mantemento_Sanidade_XXXX_XX.xlsx — hojas LAB 201/203/205/207 con planes activos

Con DRY_RUN=True solo muestra estadísticas; cambia a False para generar los archivos.
"""

import os
import sys
import shutil
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment

sys.path.insert(0, os.path.dirname(__file__))
from base import conectar, leer

# ─────────────── CONFIGURACIÓN ───────────────
CURSO_ACADEMICO  = "2025-2026"
CURSO_CORTO      = "2025_26"       # para el nombre del archivo de salida

DRY_RUN = True   # Cambia a False para generar los archivos Excel
# ──────────────────────────────────────────────

RAIZ      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATES = os.path.join(RAIZ, 'assets', 'templates')
EXPORTS   = os.path.join(RAIZ, 'exports')

TEMPLATE_INV = os.path.join(TEMPLATES, 'CIFP Manuel Antonio_Inventarios_Curso 2025-26.xlsx')
TEMPLATE_MAN = os.path.join(TEMPLATES, 'MD84MAN01_Plan_mantemento_Sanidade.xlsx')

# Mapeo de valores ubicacion → nombre de hoja del template
LAB_MAPPING = {
    'Lab 201': 'LAB 201',
    'Lab 203': 'LAB 203',
    'Lab 205': 'LAB 205',
    'Lab 207': 'LAB 207',
}

# Primer día de cada periodo del curso 2025-2026
FECHAS_CURSO = {
    'Mensual'      : date(2025, 9, 1),
    'Trimestral'   : date(2025, 9, 1),
    'Semestral'    : date(2025, 9, 1),
    'Anual'        : date(2025, 9, 1),
    'Bianual'      : date(2025, 9, 1),
    'Trianual'     : date(2025, 9, 1),
    'Pretemporada' : date(2025, 9, 1),
    'Posttemporada': date(2026, 6, 1),
}


# ─── Helpers ──────────────────────────────────

def ubicacion_to_lab(ubicacion):
    """Convierte el valor ubicacion de un equipo al nombre de hoja del template."""
    if not ubicacion:
        return None
    if ubicacion in LAB_MAPPING:
        return LAB_MAPPING[ubicacion]
    for num in ['201', '203', '205', '207']:
        if num in ubicacion:
            return f'LAB {num}'
    return None


def format_serie(val):
    """Convierte número de serie a string limpio (por si llega como número)."""
    if val is None or val == '':
        return ''
    if isinstance(val, float):
        return str(int(val))
    return str(val).strip()


def fecha_str(val):
    """Normaliza fechas a string dd/mm/yyyy; pasa strings sin tocar."""
    if not val:
        return ''
    if isinstance(val, (date, datetime)):
        return val.strftime('%d/%m/%Y')
    return str(val)


_THIN = Side(style='thin')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FONT   = Font(name='Calibri', size=10)
_ALIGN  = Alignment(wrap_text=False, vertical='center')


def estilizar(cell):
    cell.border = _BORDER
    cell.font   = _FONT
    cell.alignment = _ALIGN


def escribir_fila(ws, row_num, valores):
    for col_i, val in enumerate(valores, start=1):
        cell = ws.cell(row_num, col_i)
        cell.value = val
        estilizar(cell)


# ─── Inventario ───────────────────────────────

def generar_inventario(conn, dry_run=True):
    print('\n=== INVENTARIO DE EQUIPOS (hoja Sanidade) ===')

    _, _, equipos = leer(conn, 'equipos')

    conteo = {}
    for e in equipos:
        k = ubicacion_to_lab(e.get('ubicacion', '')) or '(sin lab)'
        conteo[k] = conteo.get(k, 0) + 1

    print(f'Total equipos: {len(equipos)}')
    for k in sorted(conteo):
        print(f'  {k}: {conteo[k]}')

    if dry_run:
        print('\n[DRY_RUN=True] No se ha generado ningún archivo.')
        return

    os.makedirs(EXPORTS, exist_ok=True)
    dest = os.path.join(EXPORTS, f'Inventarios_Sanidade_{CURSO_CORTO}.xlsx')
    shutil.copy2(TEMPLATE_INV, dest)

    wb = openpyxl.load_workbook(dest)
    ws = wb['Sanidade']

    # Cabeceras en fila 8; datos desde fila 9
    fila = 9
    equipos_sorted = sorted(equipos, key=lambda e: (
        e.get('ubicacion', '') or '',
        e.get('tipo_equipo', '') or '',
        e.get('id_activo', '') or '',
    ))

    for e in equipos_sorted:
        marca_modelo = ' '.join(filter(None, [
            str(e.get('marca') or '').strip(),
            str(e.get('modelo') or '').strip(),
        ]))
        descripcion = (e.get('observaciones') or '').strip() or (e.get('estado_operativo') or '').strip()

        escribir_fila(ws, fila, [
            f"{e.get('tipo_equipo', '')} ({e.get('id_activo', '')})",  # A — Denominación
            e.get('ubicacion', ''),                                      # B — Ubicación
            marca_modelo,                                                 # C — Marca, modelo
            format_serie(e.get('numero_serie')),                         # D — Nº de serie
            1,                                                            # E — Nº de unidades
            descripcion,                                                  # F — Descripción
        ])
        fila += 1

    wb.save(dest)
    print(f'\n✓ Guardado: {dest}')
    print(f'  {fila - 9} equipos añadidos')


# ─── Plan de mantenimiento ────────────────────

def generar_plan_mantenimiento(conn, dry_run=True):
    print('\n=== PLAN DE MANTEMENTO PREVENTIVO ===')

    _, _, equipos_list = leer(conn, 'equipos')
    _, _, planes_list  = leer(conn, 'planes_mantenimiento')
    _, _, registros    = leer(conn, 'registro_mantenimientos')

    equipos = {e['id_activo']: e for e in equipos_list}

    # Índice de registros: (id_plan, curso) → registro más reciente
    reg_idx = {}
    for r in registros:
        key = (r.get('id_plan', ''), r.get('curso_academico', ''))
        fecha = r.get('fecha_realizacion', '') or ''
        prev  = reg_idx.get(key)
        if not prev or fecha > (prev.get('fecha_realizacion') or ''):
            reg_idx[key] = r

    # Agrupar planes activos por lab
    lab_planes = {'LAB 201': [], 'LAB 203': [], 'LAB 205': [], 'LAB 207': []}
    sin_lab    = []

    for p in planes_list:
        if not p.get('activo'):
            continue
        equipo = equipos.get(p.get('id_equipo', ''))
        if not equipo:
            continue
        lab = ubicacion_to_lab(equipo.get('ubicacion', ''))
        if lab in lab_planes:
            lab_planes[lab].append((p, equipo))
        else:
            sin_lab.append((p, equipo))

    print('Planes activos por lab:')
    for lab, items in lab_planes.items():
        print(f'  {lab}: {len(items)} planes')
    if sin_lab:
        print(f'  Sin lab asignado: {len(sin_lab)} planes')
        for p, e in sin_lab[:5]:
            print(f'    {e.get("id_activo")} — {e.get("ubicacion")} — {p.get("id_plan")}')

    if dry_run:
        print('\n[DRY_RUN=True] No se ha generado ningún archivo.')
        return

    os.makedirs(EXPORTS, exist_ok=True)
    dest = os.path.join(EXPORTS, f'Plan_Mantemento_Sanidade_{CURSO_CORTO}.xlsx')
    shutil.copy2(TEMPLATE_MAN, dest)

    wb = openpyxl.load_workbook(dest)

    # Actualizar curso en portada
    portada = wb['Portada']
    for row in portada.iter_rows():
        for cell in row:
            if cell.value == '2023_24':
                cell.value = CURSO_CORTO

    for lab_name, items in lab_planes.items():
        if lab_name not in wb.sheetnames:
            print(f'  AVISO: hoja "{lab_name}" no encontrada en el template')
            continue

        ws = wb[lab_name]
        ws['A3'].value = f'Data de actualización: {datetime.now().strftime("%d/%m/%Y")}'

        # Cabeceras en fila 10; datos desde fila 11
        fila = 11
        items_sorted = sorted(items, key=lambda x: (
            x[1].get('tipo_equipo', '') or '',
            x[0].get('id_plan', '') or '',
        ))

        for plan, equipo in items_sorted:
            reg = reg_idx.get((plan.get('id_plan', ''), CURSO_ACADEMICO))

            fecha_realiz = fecha_str(reg.get('fecha_realizacion') if reg else '')
            supervisado  = ((reg.get('supervisado_por') or '') if reg else '').strip()
            obs_reg      = ((reg.get('observaciones') or '') if reg else '').strip()

            periodicidad  = plan.get('periodicidad', '')
            fecha_prevista = fecha_str(FECHAS_CURSO.get(periodicidad))

            escribir_fila(ws, fila, [
                f"{equipo.get('tipo_equipo', '')} ({equipo.get('id_activo', '')})",  # A — Denominación
                equipo.get('ubicacion', ''),                                           # B — Ubicación
                equipo.get('responsable', ''),                                         # C — Responsable
                plan.get('tipo_intervencion', ''),                                     # D — Interno/Externo
                periodicidad,                                                           # E — Periodicidade
                plan.get('operacion', ''),                                             # F — Operación
                fecha_prevista,                                                        # G — Data prevista
                fecha_realiz,                                                          # H — Data realización
                supervisado,                                                           # I — Supervisado por
                obs_reg,                                                               # J — Observacións
            ])
            fila += 1

        print(f'  {lab_name}: {fila - 11} filas')

    wb.save(dest)
    print(f'\n✓ Guardado: {dest}')


# ─── Main ─────────────────────────────────────

def main():
    conn = conectar()
    generar_inventario(conn, dry_run=DRY_RUN)
    generar_plan_mantenimiento(conn, dry_run=DRY_RUN)
    if DRY_RUN:
        print('\nPara generar los archivos: cambia DRY_RUN = False y vuelve a ejecutar.')
    conn.close()


if __name__ == '__main__':
    main()
